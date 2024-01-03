import openpyxl

#워크북 인스턴스 객체
wb=openpyxl.Workbook()
#활성화된 워크시트 객체
sheet=wb.active
#워크시트 제목
sheet.title='회원정보'

#헤더컬럼
header_titles=['아이디', '전화번호']
for idx, title in enumerate(header_titles):
    sheet.cell(row=1, column=idx+1, value=title)

#내용 저장
members=[('happy', '010-1234-5678'),('smile','010-9876-5432')]
row_num=2

for r, member in enumerate(members): #회원정보 목록 탐색
    for c, v in enumerate(member):
        sheet.cell(row=row_num+r, column=c+1, value=v)

wb.save('members.xlsx')
wb.close()