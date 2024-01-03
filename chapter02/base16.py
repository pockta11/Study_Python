import openpyxl

#부분 로딩
'''
wb=openpyxl.load_workbook('sample.xlsx')
sheet=wb['Sheet1']
print(sheet.max_column, sheet.max_row) #총 컬럼, 총 행의 수
print(sheet.cell(row=1, column=1).value)
print(sheet.cell(row=2, column=1).value)
wb.close()
'''
'''
# 행단위 로딩

wb=openpyxl.load_workbook('sample.xlsx')
sheet=wb['Sheet1']
cells=sheet['A2':'C4']
for row in cells:
    for cell in row:
        print(cell.value)
    print('-'*10)
wb.close()
'''


# 모두 로딩
wb=openpyxl.load_workbook('sample.xlsx')
sheet=wb['Sheet1']
cells=sheet['A2':'C4']

#워크시트 객체의 iter_rows() 함수를 이용해서 워크시트 내의 모든 row를 탐색
#min_row=2 : 초기 탐색 위치 지정
for row in sheet.iter_rows(min_row=2):
    for cell in row:
        print(cell.value)
    print('-'*10)
wb.close()
